# Author: Selorm Kwami Dzakpasu

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle
from datetime import datetime

def process_caiso_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['csv', 'xlsx']:
        raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")

    # Load the file into a DataFrame
    if file_extension == 'csv':
        df = pd.read_csv(file_path, header=None)
    else:
        df = pd.read_excel(file_path, header=None)

    # Remove the first three rows
    df = df.iloc[3:, :]

    # Reset the index after dropping rows
    df.reset_index(drop=True, inplace=True)

    # Set the fourth row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Reset the index again
    df.reset_index(drop=True, inplace=True)

    # Check for two consecutive empty cells in the first column
    empty_cell_count = 0
    last_filled_row = len(df) - 1  # Default to the last row if no empty cells are found

    for i in range(len(df)):
        if pd.isnull(df.iloc[i, 0]):  # Check if the cell is empty (NaN)
            empty_cell_count += 1
        else:
            empty_cell_count = 0
        
        # If two consecutive empty cells are found, break the loop
        if empty_cell_count == 2:
            last_filled_row = i - 2  # Set to the row before the first empty cell
            break

    # Slice the DataFrame to keep only the rows up to the last filled row in the first column
    df = df.iloc[:last_filled_row + 1, :]
    
    # Replace newline characters in column headers with a space
    df.columns = df.columns.str.replace('\n', ' ', regex=True)
    
    # Drop redundant columns
    df.drop(['Fuel-1', 'Fuel-2', 'Fuel-3', 'Suspension Status', 'Interconnection Request Receive Date', 'Proposed On-line Date (as filed with IR)', 'Study Process'], axis=1, inplace=True)

    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "Station or Transmission Line": "POI Name",
    "Utility": "Transmission Owner",
    "Current On-line Date": "Projected COD",
    "Interconnection Agreement  Status": "Interconnection Agreement Status",
    "Feasibility Study or Supplemental Review": "Feasibility Study Status",
    "System Impact Study or  Phase I Cluster Study": "System Impact Study Status",
    "Facilities Study (FAS) or  Phase II Cluster Study": "Facilities Study Status",
    "Net MWs to Grid": "Capacity (MW)",
    "Optional Study (OS)": "Optional Study"
    }
    
    # Rename Columns
    df.rename(columns=columns_to_rename, inplace=True)
    
    # Concatenate 'Full Capacity, Partial or Energy Only (FC/P/EO)' and 'Off-Peak Deliverability and Economic Only' columns
    df['Capacity Status'] = df[['Full Capacity, Partial or Energy Only (FC/P/EO)', 'Off-Peak Deliverability and Economic Only']
                            ].apply(lambda row: ' , '.join(row.dropna().astype(str)), axis=1)

    # Drop the original columns
    df.drop(['Full Capacity, Partial or Energy Only (FC/P/EO)', 'Off-Peak Deliverability and Economic Only'], axis=1, inplace=True)

    # Save the modified DataFrame back to a file
    output_file = f"Processed Queues/CAISO_Queue.xlsx"
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # load workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format
            cell.alignment = Alignment(horizontal="left", vertical="center") # Apply cell alignment

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/CAISO Queue Report.xlsx"  # Replace with your file name
process_caiso_file(file_path)