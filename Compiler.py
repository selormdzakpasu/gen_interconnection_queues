# Author: Selorm Kwami Dzakpasu

from openpyxl import load_workbook
import pandas as pd

# Function to add ISO name in the first column of their respective queues
def add_column(file_path, iso_name, sheet_name=None):
    # Load the workbook
    wb = load_workbook(file_path)
    
    # Default to the first sheet if no sheet name is provided
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    
    # Insert a new column at the first position
    ws.insert_cols(1)
    
    # Set the header for the new column
    ws.cell(row=1, column=1).value = "ISO"
    
    # Fill the column with the entity name up to the last filled row
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).value = iso_name
    
    # Save the workbook
    wb.save(file_path)
    print(f"Column 'ISO' added successfully to {ws.title} in {file_path}")


# Function to combine all ISO Queues
def concatenate_excel_files(file_paths, output_file):
    # List to hold DataFrames
    dataframes = []
    
    # Read each file and append the DataFrame to the list
    for file in file_paths:
        df = pd.read_excel(file)
        dataframes.append(df)
    
    # Concatenate all DataFrames
    combined_df = pd.concat(dataframes, ignore_index=True)
    
    # Reorder the DataFrame columns
    new_column_order = [
        'ISO', 'Project ID', 'Project Name', 'Queue Position', 'Transmission Owner', 'Queue Date', 'State', 
        'County', 'Service Type', 'Application Status', 'Project Status', 'POI Name', 'Projected COD', 
        'Technology', 'Type-1', 'Type-2', 'Type-3', 'Capacity (MW)', 'MW-1', 'MW-2', 'MW-3', 'Summer MW', 
        'Winter MW', 'MW Energy', 'Capacity Status', 'Interconnection Agreement Status', 
        'Approved for Synchronization', 'Actual In Service Date', 'Done Date', 'Withdrawal Date', 
        'Cessation Date', 'TPD Allocation Percentage', 'TPD Allocation Group', 'Availability of Studies', 
        'Feasibility Study Status', 'System Impact Study Status', 'Facilities Study Status', 'Screening Study Started', 
        'Screening Study Complete', 'FIS Requested', 'FIS Approved', 'Optional Study', 'Economic Study Required', 
        'PTO Study Region', 'Study Cycle', 'Study Group', 'CDR Reporting Zone', 'Current Cluster', 'Cluster Group', 
        'Comment', 'Change Indicators'
    ]
    combined_df = combined_df[new_column_order]
    
    # List of project statuses to remove (Withdrawn/Deactivated)
    proj_statuses_withdrawn = [
        "Annulled", "Canceled", "Deactivated", "Retracted", "Suspended", "WITHDRAWN", "Withdrawn"
    ]
    
    # List of project statuses to remove (In Service/IA FULLY EXECUTED/COMMERCIAL OPERATION)
    proj_statuses_completed = ["IA FULLY EXECUTED/COMMERCIAL OPERATION", "In Service"]

    # Create a DataFrame for entries with withdrawn status
    withdrawn_df_1 = combined_df[combined_df["Application Status"] == "Withdrawn"] # Check "Application Status" for withdrawn entries
    not_withdrawn_df = combined_df[combined_df["Application Status"] != "Withdrawn"]
    
    withdrawn_df_2 = not_withdrawn_df[not_withdrawn_df["Withdrawal Date"].notna()] # To filter out rows where 'Withdrawal Date' is not empty (not NaN or not empty string)
    not_withdrawn_df = not_withdrawn_df[~not_withdrawn_df["Withdrawal Date"].notna()]
    
    withdrawn_df_3 = not_withdrawn_df[not_withdrawn_df["Project Status"].isin(proj_statuses_withdrawn)] # Check "Project Status" for withdrawn entries
    
    withdrawn = [withdrawn_df_1, withdrawn_df_2, withdrawn_df_3]
    withdrawn_df = pd.concat(withdrawn, ignore_index=True)

    # Remove rows with withdrawn status from the main dataframe
    combined_df = combined_df[combined_df["Application Status"] != "Withdrawn"]
    combined_df = combined_df[~combined_df["Withdrawal Date"].notna()] # To filter out rows where 'Withdrawal Date' is empty or NaN
    combined_df = combined_df[~combined_df["Project Status"].isin(proj_statuses_withdrawn)]

    # Create a DataFrame for entries with in-service status
    completed_df = combined_df[combined_df["Project Status"].isin(proj_statuses_completed)] # Check "Project Status" for completed entries

    # Remove rows with completed status from the main dataframe
    combined_df = combined_df[~combined_df["Project Status"].isin(proj_statuses_completed)] # Remove in service entries from main dataframe

    # Export DataFrames to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name="Active") # Active entries
        withdrawn_df.to_excel(writer, index=False, sheet_name="Withdrawn") # Withdrawn entries
        completed_df.to_excel(writer, index=False, sheet_name="Completed") # Entries with "done" and "in service" status
    
    print(f"Files have been concatenated into {output_file}")

# File paths and corresponding ISO names
files_and_iso = [
    ("Processed Queues/CAISO_Queue.xlsx", "CAISO"),
    ("Processed Queues/ERCOT_Queue.xlsx", "ERCOT"),
    ("Processed Queues/MISO_Queue.xlsx", "MISO"),
    ("Processed Queues/PJM_Queue.xlsx", "PJM"),
    ("Processed Queues/SPP_Queue.xlsx", "SPP"),
    ("Processed Queues/NEISO_Queue.xlsx", "NEISO"),
    ("Processed Queues/NYISO_Queue.xlsx", "NYISO"),
]

# File paths for concatenation
file_paths = [
    "Processed Queues/CAISO_Queue.xlsx", "Processed Queues/ERCOT_Queue.xlsx", "Processed Queues/MISO_Queue.xlsx",
    "Processed Queues/PJM_Queue.xlsx", "Processed Queues/SPP_Queue.xlsx", "Processed Queues/NEISO_Queue.xlsx", "Processed Queues/NYISO_Queue.xlsx"
]

# Process each file, filling the first column with the ISO name
for file_path, iso_name in files_and_iso:
    add_column(file_path, iso_name)


output_file = "Combined_Queues.xlsx" # "Combined_Queues.xlsx" can be replaced with the desired file path & name
concatenate_excel_files(file_paths, output_file)
