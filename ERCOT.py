# Author: Selorm Kwami Dzakpasu

# File extension should be .xlsx.
# .csv not supported

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def process_ercot_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['xlsx']:
        raise ValueError("Unsupported file format. Please provide a .xlsx file.")
    
    # Load the Excel workbook
    workbook = load_workbook(filename=file_path, data_only=True)

    # Helper function to process individual sheets
    def process_sheet(sheet_name, remove_rows, header_merge_range):
        sheet = workbook[sheet_name]
        
        # Remove the first rows
        for _ in range(remove_rows):
            sheet.delete_rows(1)
        
        # Merge specified rows to create headers
        header_values = []
        for row in sheet.iter_rows(min_row=header_merge_range[0], max_row=header_merge_range[1]):
            header_values.append([str(cell.value) if cell.value else '' for cell in row])
        merged_header = [' '.join(col).strip() for col in zip(*header_values)]
        
        # Remove merged header rows from the sheet
        for _ in range(header_merge_range[1] - header_merge_range[0] + 1):
            sheet.delete_rows(1)
        
        # Convert sheet to DataFrame
        data = pd.DataFrame(sheet.values)
        
        # Set merged header as column names
        data.columns = merged_header
        
        # Drop empty rows or columns (if any)
        data.dropna(how='all', inplace=True)
        
        return data

    # Process Worksheet 5
    sheet5_data = process_sheet(
        sheet_name=workbook.sheetnames[4],  # Zero-based indexing for sheets
        remove_rows=30,
        header_merge_range=(1, 5),
    )
    
    # Process Worksheet 6
    sheet6_data = process_sheet(
        sheet_name=workbook.sheetnames[5],
        remove_rows=14,
        header_merge_range=(1, 3),
    )
    
    # Merge the two datasets
    merged_data = pd.concat([sheet5_data, sheet6_data], ignore_index=True)

    # Save the combined data to a new Excel file
    output_file = f"Processed Queues/ERCOT_Queue.xlsx"
    merged_data.to_excel(output_file, index=False)

    # Re-open the newly saved file to apply the column renaming and merging
    wb = load_workbook(output_file)
    ws = wb.active

    # Load the data into a DataFrame
    processed_df = pd.read_excel(output_file)
    
    # Merge Change Indicator columns
    processed_df['Change Indicators'] = processed_df[['Change indicators: Proj Name, MW Size, COD, SFS/NtP, FIS Request, Status Change INA-to-PLN', 
                                                      'Change indicators: Proj Name, MW Size, COD, SFS. Status Change INA-to-PLN']].apply(lambda row: ' '.join(row.dropna().astype(str)), axis=1)
    
    # Drop redundant columns
    processed_df.drop(['Fuel', 'Change indicators: Proj Name, MW Size, COD, SFS/NtP, FIS Request, Status Change INA-to-PLN',
                       'Change indicators: Proj Name, MW Size, COD, SFS. Status Change INA-to-PLN', 'Approval Date for Submission of Proof of  Site Control',
                       'Air Permit', 'GHG Permit', 'Water Availability', 'Construction Start', 'Construction End', 'Approved for Energization',
                       'Financial Security  Required to fund Dist. Upgrades', 'Meets Planning Guide Section 6.9(1)  Requirements for  Inclusion in Planning  Models',
                       'Meets All Planning Guide Section 6.9 Requirements for  Inclusion in Planning Models', 'Meets Planning Guide  QSA (Section 5.9)  Prerequisites',
                       'Financial Security  and Notice to  Proceed Provided', 'Model Ready Date'], axis=1, inplace=True)
    
    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "INR": "Project ID",
    "Interconnecting Entity": "Transmission Owner",
    "POI Location": "POI Name",
    "IA Signed": "Queue Date",
    "GIM Study Phase": "Availability of Studies"
    }
    
    # Rename Columns
    processed_df.rename(columns=columns_to_rename, inplace=True)

    # Save the modified DataFrame back to the same file
    processed_df.to_excel(output_file, index=False)
    
    # Load the workbook and apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/ERCOT GIS_Report_December2024.xlsx"  # Replace with your file name
process_ercot_file(file_path)