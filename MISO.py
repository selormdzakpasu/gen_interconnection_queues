# Author: Selorm Kwami Dzakpasu

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def process_miso_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['csv', 'xlsx']:
        raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")

    # Load the file into a DataFrame
    if file_extension == 'csv':
        df = pd.read_csv(file_path, header=None)
    else:
        df = pd.read_excel(file_path, header=None)

    # Set the first row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Reset the index
    df.reset_index(drop=True, inplace=True)
    
    # Remove row that contains "xylophone"
    df = df[~df.apply(lambda row: row.astype(str).str.contains('xylophone', case=False).any(), axis=1)]
    
    # Drop redundant columns
    df.drop(['Negotiated In Service Date', 'Fuel', 'Decision Point 1 ERIS MW', 'Decision Point 1 NRIS MW', 'Decision Point 2 ERIS MW',
             'Decision Point 2 NRIS MW'], axis=1, inplace=True)
    
    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "Project #": "Project ID",
    "Request Status": "Application Status",
    "Withdrawn Date": "Withdrawal Date",
    "Generating Facility": "Technology",
    "Appl In Service Date": "Projected COD",
    "Post GIA Status": "Project Status",
    "Study Phase": "Availability of Studies"
    }
    
    # Rename Columns
    df.rename(columns=columns_to_rename, inplace=True)

    # Save the modified DataFrame back to a file
    output_file = f"Processed Queues/MISO_Queue.xlsx"
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')
    
        # Load Workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")
    
    date_columns = ['Queue Date', 'Withdrawal Date', 'Done Date', 'Projected COD']

    # Find the column indices for the date columns
    date_column_indices = [cell.column for cell in ws[1] if cell.value in date_columns]

    # Iterate over rows and only process the specified "Date" columns
    for row in ws.iter_rows(min_row=2):  # Start from row 2 to avoid the header row
        for col_idx in date_column_indices:
            cell = row[col_idx - 1]  # Adjust for 0-based index
            if isinstance(cell.value, str) and 'T' in cell.value:
                cell.value = cell.value.split('T')[0]  # Modify the cell value
            elif isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/MISO GI Interactive Queue.csv"  # Replace with your file name
process_miso_file(file_path)