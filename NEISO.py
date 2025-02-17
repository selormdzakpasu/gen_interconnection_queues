# Author: Selorm Kwami Dzakpasu

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle
from datetime import datetime

def process_neiso_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['csv', 'xlsx']:
        raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")

    # Load the file into a DataFrame
    if file_extension == 'csv':
        df = pd.read_csv(file_path, header=None)
    else:
        df = pd.read_excel(file_path, header=None)

    # Remove the first four rows
    df = df.iloc[4:, :]

    # Reset the index after dropping rows
    df.reset_index(drop=True, inplace=True)

    # Set the fifth row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Reset the index again
    df.reset_index(drop=True, inplace=True)
    
    # Drop redundant columns
    df.drop(['Updated', 'Unit', 'Sync Date', 'Serv', 'SIS Complete', 'I39', 'TO Report', 'Dev', 'Zone'], axis=1, inplace=True)

    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "Position": "Queue Position",
    "Type": "Service Type",
    "Requested": "Queue Date",
    "Alternative Name": "Project Name",
    "Fuel Type": "Technology",
    "Net MW": "Capacity (MW)",
    "Op Date": "Projected COD",
    "W/ D Date": "Withdrawal Date",
    "Interconnection Location": "POI Name",
    "FS": "Feasibility Study Status",
    "SIS": "System Impact Study Status",
    "OS": "Optional Study",
    "FAC": "Facilities Study Status",
    "IA": "Interconnection Agreement Status"
    }
    
    # Rename Columns
    df.rename(columns=columns_to_rename, inplace=True)

    # Save the modified DataFrame back to a file
    output_file = f"Processed Queues/NEISO_Queue.xlsx"
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format
            cell.alignment = Alignment(horizontal="left", vertical="center") # Apply alignment to cells

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/NEISO QueueReport.xlsx"  # Replace with your file name
process_neiso_file(file_path)