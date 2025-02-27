# Author: Selorm Kwami Dzakpasu

# File extension should be .xlsx.
# .csv not supported

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def process_nyiso_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['xlsx']:
        raise ValueError("Unsupported file format. Please provide a .xlsx file.")
    
    # Load the file into a DataFrame
    df = pd.read_excel(file_path, header=None)

    # Check for an empty cell in the first column to determine the last filled row
    empty_cell_count = 0
    last_filled_row = len(df) - 1  # Default to the last row if no empty cells are found

    for i in range(len(df)):
        if pd.isnull(df.iloc[i, 0]):  # Check if the cell is empty (NaN)
            empty_cell_count += 1
        else:
            empty_cell_count = 0
        
        # If two consecutive empty cells are found, break the loop
        if empty_cell_count == 1:
            last_filled_row = i - 1  # Set to the row before the first empty cell
            break

    # Slice the DataFrame to keep only the rows up to the last filled row using the first column
    df = df.iloc[:last_filled_row + 1, :]
    
    # Set the first row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Reset the index
    df.reset_index(drop=True, inplace=True)
    
    # Merge "Developer/Interconnection Customer" and "Utility"
    df['Transmission Owner'] = df[['Developer/Interconnection Customer', 'Utility']].apply(lambda row: ' , '.join(row.dropna().astype(str)), axis=1)
     
    # Drop redundant columns
    df.drop(['Z', 'Last Updated Date', 'IA Tender Date', 'Proposed In-Service/Initial Backfeed Date', 'Proposed Sync Date', 'CY/FS Complete Date',
             'Developer/Interconnection Customer', 'Utility'], axis=1, inplace=True)

    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "Queue Pos.": "Queue Position",
    "Date of IR": "Queue Date",
    "SP (MW)": "Summer MW",
    "WP (MW)": "Winter MW",
    "Type/ Fuel": "Technology",
    "Points of Interconnection": "POI Name",
    "S": "Project Status",
    "Proposed COD": "Projected COD"
    }
    
    # Rename Columns
    df.rename(columns=columns_to_rename, inplace=True)
    
    # Define the mapping key for the "Technology" column
    technology_key = {
        "ST": "Steam Turbine",
        "CT": "Combustion Turbine",
        "CS": "Steam Turbine & Combustion Turbine",
        "H": "Hydro",
        "PS": "Pumped Storage",
        "W": "Wind",
        "OSW": "Off-Shore Wind",
        "NU": "Nuclear",
        "NG": "Natural Gas",
        "M": "Methane",
        "SW": "Solid Waste",
        "S": "Solar",
        "Wo": "Wood",
        "ES": "Energy Storage",
        "O": "Oil",
        "C": "Coal",
        "D": "Dual Fuel",
        "L": "Load",
        "FC": "Fuel Cell",
        "CW": "Energy Storage + Wind",
        "CR": "Energy Storage + Solar"
    }
    
    # Define the mapping key for the "Project Status" column
    project_status_key = {
        "1": "Scoping Meeting Pending",
        "2": "FES Pending",
        "3": "FES in Progress",
        "3A": "FES Approved/Performed",
        "4": "SRIS/SIS Pending",
        "5": "SRIS/SIS in Progress",
        "5P": "SRIS Commenced, Stopped and Pending Adoption of IP",
        "6": "SRIS/SIS Approved",
        "7": "FS Pending",
        "8": "Rejected Cost Allocation/Next FS Pending",
        "9": "FS in Progress",
        "10": "Accepted Cost Allocation/IA in Progress",
        "11": "IA Completed",
        "12": "Under Construction",
        "13": "In Service for Test",
        "14": "In Service Commercial",
        "0": "Withdrawn",
        "15": "Partial In-Service",
        "P": "Pending Adoption of IP Compliance with Order 2023"
    }
    
    # Replace the values in column index 6 based on the technology key
    df['Technology'] = df['Technology'].replace(technology_key)
    
    # Replace values in column index 11 based on the project status key
    df['Project Status'] = df['Project Status'].replace(project_status_key)

    # Save the modified DataFrame back to a file
    output_file = f"Processed Queues/NYISO_Queue.xlsx"
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load Workbook
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/NYISO-Interconnection-Queue.xlsx"  # Replace with your file name
process_nyiso_file(file_path)