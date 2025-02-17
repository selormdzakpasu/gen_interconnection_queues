# Author: Selorm Kwami Dzakpasu

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def process_spp_file(file_path):
    # Check the file extension
    file_extension = file_path.split('.')[-1].lower()

    if file_extension not in ['csv', 'xlsx']:
        raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")

    # Load the file into a DataFrame
    if file_extension == 'csv':
        df = pd.read_csv(file_path, header=None)
    else:
        df = pd.read_excel(file_path, header=None)
   
    # Remove the first row
    df = df.iloc[1:, :]

    # Reset the index after dropping row
    df.reset_index(drop=True, inplace=True)
    
    # Set the second row as column headers
    df.columns = df.iloc[0]
    df = df[1:]

    # Reset the index again
    df.reset_index(drop=True, inplace=True)
    
    # List of columns to drop
    columns_to_drop = ['IFS Queue Number', 'In-Service Date', 'Fuel Type', 'Cause of Delay']

    # Check if 'Replacement Generator Commercial Op Date' or 'Original Generator Commercial Op Date' exists and add to the list
    if 'Replacement Generator Commercial Op Date' in df.columns:
        columns_to_drop.append('Replacement Generator Commercial Op Date')
    elif 'Original Generator Commercial Op Date' in df.columns:
        columns_to_drop.append('Original Generator Commercial Op Date')

    # Drop the columns
    df.drop(columns=columns_to_drop, axis=1, inplace=True)

    # Rename specific columns based on the provided mapping
    columns_to_rename = {
    "Generation Interconnection Number": "Project ID",
    " Nearest Town or County": "County",
    "TO at POI": "Transmission Owner",
    "Capacity": "Capacity (MW)",
    "Commercial Operation Date": "Projected COD",
    "MAX Summer MW": "Summer MW",
    "MAX Winter MW": "Winter MW",
    "Generation Type": "Technology",
    "Substation or Line": "POI Name",
    "Request Received": "Queue Date",
    "Date Withdrawn": "Withdrawal Date",
    "Status": "Project Status"
    }
    
    # Rename Columns
    df.rename(columns=columns_to_rename, inplace=True)

    # Save the modified DataFrame back to a file
    output_file = f"Processed Queues/SPP_Queue.xlsx"
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load workbook
    wb = load_workbook(output_file)
    ws = wb.active

    # Create a date format style
    date_style = NamedStyle(name="short_date", number_format="MM/DD/YYYY")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):  # Check if the cell contains a date
                cell.style = date_style  # Apply short date format

    wb.save(output_file)

    print(f"File processed and saved at {output_file}")

# Specify file name and run
file_path = "Queues/SPP GI_ActiveRequest.xlsx"  # Replace with your file name
process_spp_file(file_path)